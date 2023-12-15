import asyncio
import datetime
import os
import comtypes.client
import openpyxl
import pandas as pd
import pyppeteer.errors
from bs4 import BeautifulSoup
from openpyxl.styles import Border, Side, Alignment, Font
from pyppeteer import launch

user = "p_1galmeida"
password = "allgood19"


async def main():
    # Gets current date's ISOWEEK so I can access only the associated week's class. So it doesn't have to access
    # all of them runtime. It also adjusts them to account for first and second semester. The math looks weird but works
    current_week = datetime.date.isocalendar(datetime.date.today())[1]
    if current_week > 6:
        if current_week > 26:
            week_index = current_week - 32
        else:
            week_index = current_week - 6
    else:
        week_index = 0

    # Launches an internet browser and accesses the main page for login.
    browser = await launch()
    page = await browser.newPage()

    await page.setUserAgent(
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 '
        'Safari/537.36')

    page_path = 'https://www.avaeduc.com.br'
    await page.goto(page_path, {'waitUntil': 'load', 'timeout': 0})

    # Logs into page with my credentials and clicks the submit button. Then it sleeps for 10 seconds so the page loads.
    # Couldn't make {'waitUntil': 'load'} work on this. Need to figure it out so I don't hardcode the time.
    await page.type('#units', 'Unic')
    await page.type('[id = username]', user)
    await page.type('[name = password]', password)
    await page.click('[type = submit]')
    print('Accessing AVA')
    await asyncio.sleep(10)
    try:
        await page.click('[id = drawer-toggle-button]')
    except pyppeteer.errors.PageError:
        print('Timed out. Trying again.')
        await asyncio.sleep(20)
        await page.click('[id = drawer-toggle-button]')
    # Gets page html content for bs4 parsing, which allows me to get the links for the courses I teach.
    # Then, from this html and links, I get the unit I need to access and ask the program to access them.
    # The last one is a website side bug, so I just have it removed.
    # The bug is fixed as of 2023-1, thus, the line is commented out.

    html = await page.content()
    soup = BeautifulSoup(html, 'html.parser')
    courses = soup.select('[data-key]')
    keys = [course.get('data-key', None) for course in courses if course.get('data-key', None).isnumeric()]
    # keys.remove(keys[-1])

    counter = 1
    units = {}

    for key in keys:

        await page.goto(f'https://www.avaeduc.com.br/course/view.php?id={key}', {'waitUntil': 'load', 'timeout': 0})
        unit_soup = BeautifulSoup(await page.content(), 'html.parser')
        unit = unit_soup.select('.timeline-menu > ul > li > a')
        unit_urls = [section.get('href', None) for section in unit]
        subject_selector = unit_soup.select('#page-navbar > div > nav > ol > li:nth-child(3) > a')
        subject_name = subject_selector[0].get('title', None)
        print(f'Accessing: {subject_name} ({key})')
        units[f'subject{counter}'] = [url for url in unit_urls]

        for course_to_access in units:
            new_page = await browser.newPage()
            week_link = ""

            # Some courses don't have as many links. Estágios, mainly. Thus, I told the program to try and run the
            # current week index associated link, if that is not possible, to access the last link in course list.
            try:
                week_link = units[course_to_access][week_index]

            except IndexError:
                week_link = units[course_to_access][-1]
                print(f'Failed to access unit. Defaulted to link {week_link}')

            finally:
                await new_page.goto(week_link, {'waitUntil': 'load', 'timeout': 0})

        print(f'Accessed {subject_name} successfully!')

        counter += 1

    print('All your pages have been accessed! Enjoy your week. :)')
    await download_attendance(browser)

    year = datetime.datetime.now().year
    month = datetime.datetime.now().month
    if month > 5:
        month_tag = "2"
    else:
        month_tag = "1"

    directory = f'G:/Meu Drive/Kroton/Aulas/{year}-{month_tag}/# Listas de Presença'
    clean_spreadsheet(directory)
    convert_xlsx_to_pdf(directory)

    # This is here to keep the command line open so I can see that it is done.
    input('Press enter to exit.')


# Here I'll implement the solution to download attendance sheets for me.
async def download_attendance(browser):
    # Accesses the Participant list url
    attendance_page = await browser.newPage()
    await attendance_page.setUserAgent(
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 '
        'Safari/537.36')
    attendance_path = "https://www.avaeduc.com.br/blocks/list_participants/view.php?id=1"
    print('Waiting for Lista de participantes to load.')
    await attendance_page.goto(attendance_path, {'waitUntil': 'load', 'timeout': 0})
    print('Loaded.')

    await asyncio.sleep(3)

    await attendance_page.select('#id_institution', '35401')
    await attendance_page.select('#id_role_filter', '1')
    await asyncio.sleep(5)
    attendance_soup = BeautifulSoup(await attendance_page.content(), 'html.parser')
    collected_values = attendance_soup.select('#id_course > option')
    values = [value.get('value') for value in collected_values]
    for x in values[1:]:
        await attendance_page.select('#id_course', x)
        await asyncio.sleep(5)
        collected_subjects = BeautifulSoup(await attendance_page.content(), 'html.parser').select(
            '#id_discipline > option')
        subjects = [subject.get('value') for subject in collected_subjects]
        for i in subjects[1:]:
            print(f'Downloading attendance for group {i}, Course: {x}')
            await attendance_page.select('#id_discipline', i)
            await asyncio.sleep(5)
            await attendance_page.click('#id_submitbutton')
            await asyncio.sleep(5)

            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            if month > 5:
                month_tag = "2"
            else:
                month_tag = "1"

            content = await attendance_page.content()
            table = pd.read_html(content)
            df = pd.DataFrame(table[0])
            df.to_excel(
                f'G:/Meu Drive/Kroton/Aulas/{year}-{month_tag}/# Listas de Presença/Lista {x} {i}.xlsx',
                f'Lista {datetime.date.today()}')


def convert_xlsx_to_pdf(folder_path):
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') and not file.startswith("~"):
            print(f'Converting {file}')
            file_path = os.path.join(folder_path, file)

            xlTypePDF = 0  # Corresponding constant for PDF format in Excel

            in_file = os.path.abspath(file_path)
            out_file = os.path.abspath(os.path.join(folder_path, file.title()[0:-5]) + '.pdf')

            excel = comtypes.client.CreateObject('Excel.Application')
            wb = excel.Workbooks.Open(in_file)

            # Set print settings to fit all columns on one page and adjust margins
            ws = wb.Worksheets(1)
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False
            ws.PageSetup.LeftMargin = excel.InchesToPoints(0.25)  # Adjust margin as needed
            ws.PageSetup.RightMargin = excel.InchesToPoints(0.25)  # Adjust margin as needed

            wb.ExportAsFixedFormat(xlTypePDF, out_file)
            wb.Close(False)
            excel.Quit()


def clean_spreadsheet(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') and not filename.startswith("~"):
            print(f'Editing {filename}')
            file_path = os.path.join(directory, filename)
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.worksheets[0]

                for row in range(sheet.max_row):
                    if row != 0:
                        sheet[f"A{row+1}"] = row

                sheet.move_range(f'A1:Z{sheet.max_row}', rows=2)
                sheet['A1'] = f"{sheet['F4'].value} - {sheet['G4'].value}"
                sheet['A2'] = sheet['E4'].value[:-4]
                sheet.column_dimensions['B'].width = 45
                sheet.delete_cols(2)
                sheet.delete_cols(3, 7)



                sheet.merge_cells('A1:G1')
                sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['A1'].font = Font(bold=True)
                sheet.merge_cells('A2:G2')
                sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['A2'].font = Font(bold=True)
                sheet.merge_cells('H1:I1')
                sheet.merge_cells('H2:I2')
                sheet['H1'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['H1'].font = Font(bold=True)
                sheet['H2'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['H2'].font = Font(bold=True)
                sheet['C3'].font = Font(bold=True)

                row_counter = sheet.max_row
                for value in range(3, row_counter + 1):
                    sheet.merge_cells(f'C{value}:I{value}')
                    sheet[f'C{value}'].alignment = Alignment(horizontal='center', vertical='center')

                border_style = Border(left=Side(border_style='thin', color='000000'),
                                      right=Side(border_style='thin', color='000000'),
                                      top=Side(border_style='thin', color='000000'),
                                      bottom=Side(border_style='thin', color='000000'))

                for row in sheet[f'A1:I{sheet.max_row}']:
                    for cell in row:
                        cell.border = border_style

                sheet['B3'] = 'Nome completo'

                sheet['C3'] = 'Assinatura'
                sheet['H1'] = 'Emitido em:'
                sheet['H2'] = str(datetime.date.today().strftime('%d/%m/%Y'))

                for col in ['A', 'C', 'D', 'E']:
                    sheet.column_dimensions[col].width = 3

                workbook.save(f"{directory}/Lista {sheet['A1'].value}.xlsx")
                os.remove(file_path)

            except TypeError:
                pass

            print('Success.')

    print('All attendance sheets converted. Well done.')


asyncio.get_event_loop().run_until_complete(main())
